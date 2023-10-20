from Ecuaciones import *

# Materiales
fc = 21
fy = 420
b = 100
h = 30
fiv = 0.75
pminv =  0.0012
pminh =  0.0015
r = 5
d = h-r

# Datos iniciales refuerzo vertical
Mpy = 2009
Mny = 2704
Nbv = 4

# Datos iniciales refuerzo horizontal
Mpx = 1347
Mnx = 3041
Nbh = 4

# Datos iniciales diseño cortante simple
Vux = 158
Vuy = 157

# Acero minimo
Asminv = f_Asmin(pminv, b, h); print("Asmin = %.2f" %Asminv)
Asminh = f_Asmin(pminh, b, h); print("Asmin = %.2f" %Asminh)

#Diseño refuerzo vertical.
print(''); print('Refuerzo vertical.')
Muv = max (Mpy, Mny);                      print(" Mu  = %.1f" %Muv)
pcalv = f_pcal(fc, fy, Muv, b, d);         print(" pcal  = %.4f" %pcalv)
Ascalv = f_Ascal(pcalv, b, d);             print(" Ascal = %.2f" %Ascalv)
Asreqv = f_Asreq(Ascalv, Asminv);          print(" Asreq = %.2f" %Asreqv)
Asbv = f_Asb(Nbv);                         print(" Asb   = %.2f" %Asbv)
Sepv = f_sep(Asreqv, Asbv, b);             print(" Sep   = %.2f" %Sepv)
dbpv = f_dbp(Nbv);                         print(' db en pulg', dbpv)
Disv = f_dist(Asreqv, Asbv, b, dbpv);      print(' Ref   = ', Disv)
Ascolv = f_Ascol(Asreqv, Asbv);            print(" Ascol = %.2f" %Ascolv)
Mnv = f_Mn(fc, fy, Ascolv, b, d);          print(" Mn    = %.1f" %Mnv)
ChMnv = f_ChMn(Muv,Mnv);                   print(' Chequeo Mn =', ChMnv)

#Diseño refuerzo horizontal.
print(''); print('Refuerzo Horizontal')
Muh = max (Mpx, Mnx);                     print(" Mu  = %.1f" %Muh)
pcalh = f_pcal(fc, fy, Muh, b, d);        print(" pcal  = %.4f" %pcalh)
Ascalh = f_Ascal(pcalh, b, d);            print(" Ascal = %.2f" %Ascalh)
Asreqh = f_Asreq(Ascalh, Asminh);         print(" Asreq = %.2f" %Asreqh)
Asbh = f_Asb(Nbh);                        print(" Asb   = %.2f" %Asbh)
Seph = f_sep(Asreqh, Asbh, b);            print(" Sep   = %.2f" %Seph)
dbph = f_dbp(Nbh);                        print(' db en pulg', dbph)
Dish = f_dist(Asreqh, Asbh, b, dbph);     print(' Ref   = ', Dish)
Ascolh = f_Ascol(Asreqh, Asbh);           print(" Ascol = %.2f" %Ascolh)
Mnh = f_Mn(fc, fy, Ascolh, b, d);         print(" Mn    = %.1f" %Mnh)
ChMnh = f_ChMn(Muh,Mnh);                  print(' Chequeo Mn =', ChMnh)

#Diseño a cortante
print(''); print('Diseño a cortante')

Vcv = f_Vc(fiv, fc, b, d);                  print(" Vcv = %.1f" %Vcv)
ChVcv = f_ChVc(Vuy,Vcv);                    print(' Chequeo cortante = ', ChVcv)

Vch = f_Vc(fiv, fc, b, d);                  print(" Vch = %.1f" %Vch)
ChVch = f_ChVc(Vux,Vch);                    print(' Chequeo cortante = ', ChVch)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
book = Workbook()
book = load_workbook('D:\IEB_MUROS\PlantillaWallDesign.xlsx')
sheet = book.active

 # Datos iniciales
sheet['D7'] = b
sheet['C35'] = b*10 
sheet['D35'] = b*10
sheet['D5'] = h
    
sheet['D6'] = d; 
sheet['C34'] = d*10 
sheet['D34'] = d*10
sheet['C36'] = Vuy
sheet['D36'] = Vux   

# Acero minimo
Asminv = f_Asmin(pminv, b, h);              sheet['D13'] = "%.2f" %Asminv
Asminh = f_Asmin(pminh, b, h);              sheet['D25'] = "%.2f" %Asminh

#Diseño refuerzo vertical.
print(''); print('Refuerzo vertical.')
Muv = max (Mpy, Mny);                       sheet['D10'] = "%.1f" %Muv                      
pcalv = f_pcal(fc, fy, Muv, b, d);          sheet['D11'] = "%.4f" %pcalv
Ascalv = f_Ascal(pcalv, b, d);              sheet['D12'] = "%.2f" %Ascalv
Asreqv = f_Asreq(Ascalv, Asminv);           sheet['D14'] = "%.2f" %Asreqv
Asbv = f_Asb(Nbv);                         
Sepv = f_sep(Asreqv, Asbv, b);              sheet['D16'] = "%.1f" %Sepv
dbpv = f_dbp(Nbv);                          sheet['D15'] = dbpv                        
Disv = f_dist(Asreqv, Asbv, b, dbpv);      
Ascolv = f_Ascol(Asreqv, Asbv);            
Mnv = f_Mn(fc, fy, Ascolv, b, d);           sheet['D18'] = "%.1f" %Mnv          
ChMnv = f_ChMn(Muv,Mnv);                    sheet['D19'] = ChMnv

#Diseño refuerzo horizontal.
print(''); print('Refuerzo Horizontal')
Muh = max (Mpx, Mnx);                       sheet['D22'] = "%.1f" %Muh
pcalh = f_pcal(fc, fy, Muh, b, d);          sheet['D23'] = "%.4f" %pcalh
Ascalh = f_Ascal(pcalh, b, d);              sheet['D24'] = "%.2f" %Ascalh
Asreqh = f_Asreq(Ascalh, Asminh);           sheet['D26'] = "%.2f" %Asreqh
Asbh = f_Asb(Nbh);                        
Seph = f_sep(Asreqh, Asbh, b);              sheet['D28'] = "%.1f" %Seph
dbph = f_dbp(Nbh);                          sheet['D27'] = dbph 
Dish = f_dist(Asreqh, Asbh, b, dbph);     
Ascolh = f_Ascol(Asreqh, Asbh);           
Mnh = f_Mn(fc, fy, Ascolh, b, d);           sheet['D30'] = "%.1f" %Mnh 
ChMnh = f_ChMn(Muh,Mnh);                    sheet['D31'] = ChMnh

#Diseño a cortante
print(''); print('Diseño a cortante')

Vcv = f_Vc(fiv, fc, b, d);                   sheet['C37'] = Vcv                  
ChVcv = f_ChVc(Vux,Vcv);                     sheet['C38'] = ChVcv 

Vch = f_Vc(fiv, fc, b, d);                   sheet['D37'] = Vch 
ChVch = f_ChVc(Vux,Vch);                     sheet['D38'] = ChVch 
    

book.save ('D:\IEB_REPORTES\Wall Design.xlsx')