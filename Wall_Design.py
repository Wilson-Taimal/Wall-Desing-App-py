from Ecuaciones import *

# Función calcular
def Calcular():
    # Materiales
    fc = float (c1ent01.get())
    fy = float (c1ent02.get())
    b = float (c1ent03.get())
    h = float (c1ent04.get())
    r = float (c1ent05.get())
    fiv = float (c1ent06.get())
    pminv =  float (c1ent07.get())
    pminh =  float (c1ent08.get())
    d = h-r

    # Datos iniciales refuerzo vertical
    Mpy = float (c6ent01.get())
    Mny = float (c6ent02.get())
    Nbv = float (c2ent08.get())

    # Datos iniciales refuerzo horizontal
    Mpx = float (c6ent03.get())
    Mnx = float (c6ent04.get())
    Nbh = float (c3ent08.get())

    # Datos iniciales diseño cortante simple
    Vuy = float (c6ent05.get())
    Vux = float (c6ent06.get())
 
    # Acero minimo
    Asminv = f_Asmin(pminv, b, h); print("Asmin = %.2f" %Asminv)
    c1ent09.delete(0, 'end');    c1ent09.insert (0, "{:.2f}".format(Asminv))

    Asminh = f_Asmin(pminh, b, h); print("Asmin = %.2f" %Asminh)
    c1ent10.delete(0, 'end');    c1ent10.insert (0, "{:.2f}".format(Asminh))

    #Diseño refuerzo vertical.
    print(''); print('Refuerzo vertical.')
    Muv = max (Mpy, Mny);                      print(" Mu  = %.1f" %Muv)
    c2ent04.delete(0, 'end');                  c2ent04.insert (0, "{:.0f}".format(Muv))

    pcalv = f_pcal(fc, fy, Muv, b, d);         print(" pcal  = %.4f" %pcalv)
    c2ent05.delete(0, 'end');                  c2ent05.insert (0, "{:.4f}".format(pcalv))

    Ascalv = f_Ascal(pcalv, b, d);             print(" Ascal = %.2f" %Ascalv)
    c2ent06.delete(0, 'end');                   c2ent06.insert (0, "{:.2f}".format(Ascalv))

    Asreqv = f_Asreq(Ascalv, Asminv);          print(" Asreq = %.2f" %Asreqv)
    c2ent07.delete(0, 'end');                   c2ent07.insert (0, "{:.2f}".format(Asreqv))

    Asbv = f_Asb(Nbv);                         print(" Asb   = %.2f" %Asbv)
    Sepv = f_sep(Asreqv, Asbv, b);             print(" Sep   = %.2f" %Sepv)
    dbpv = f_dbp(Nbv);                         print(' db en pulg', dbpv)

    Disv = f_dist(Asreqv, Asbv, b, dbpv);      print(' Ref   = ', Disv)
    c2ent09.delete(0, 'end');                   c2ent09.insert (0, Disv)

    Ascolv = f_Ascol(Asreqv, Asbv);            print(" Ascol = %.2f" %Ascolv)
    c2ent13.delete(0, 'end');                   c2ent13.insert (0, "{:.2f}".format(Ascolv))

    Mnv = f_Mn(fc, fy, Ascolv, b, d);          print(" Mn    = %.1f" %Mnv)
    c2ent14.delete(0, 'end');                   c2ent14.insert (0, "{:.0f}".format(Mnv))

    ChMnv = f_ChMn(Muv,Mnv);                   print(' Chequeo Mn =', ChMnv)
    c2ent15.delete(0, 'end');                   c2ent15.insert (0, ChMnv)    

    #Diseño refuerzo horizontal.
    print(''); print('Refuerzo Horizontal')
    Muh = max (Mpx, Mnx);                     print(" Mu  = %.1f" %Muh)
    c3ent04.delete(0, 'end');                 c3ent04.insert (0, "{:.0f}".format(Muh))
    
    pcalh = f_pcal(fc, fy, Muh, b, d);        print(" pcal  = %.4f" %pcalh)
    c3ent05.delete(0, 'end');                 c3ent05.insert (0, "{:.4f}".format(pcalh))

    Ascalh = f_Ascal(pcalh, b, d);            print(" Ascal = %.2f" %Ascalh)
    c3ent06.delete(0, 'end');                 c3ent06.insert (0, "{:.2f}".format(Ascalh))

    Asreqh = f_Asreq(Ascalh, Asminh);         print(" Asreq = %.2f" %Asreqh)
    c3ent07.delete(0, 'end');                 c3ent07.insert (0, "{:.2f}".format(Asreqh))

    Asbh = f_Asb(Nbh);                        print(" Asb   = %.2f" %Asbh)
    Seph = f_sep(Asreqh, Asbh, b);            print(" Sep   = %.2f" %Seph)
    dbph = f_dbp(Nbh);                        print(' db en pulg', dbph)

    Dish = f_dist(Asreqh, Asbh, b, dbph);     print(' Ref   = ', Dish)
    c3ent09.delete(0, 'end');                 c3ent09.insert (0, Dish)

    Ascolh = f_Ascol(Asreqh, Asbh);           print(" Ascol = %.2f" %Ascolh)
    c3ent13.delete(0, 'end');                 c3ent13.insert (0, "{:.2f}".format(Ascolh))

    Mnh = f_Mn(fc, fy, Ascolh, b, d);         print(" Mn    = %.1f" %Mnh)
    c3ent14.delete(0, 'end');                 c3ent14.insert (0, "{:.0f}".format(Mnh))

    ChMnh = f_ChMn(Muh,Mnh);                  print(' Chequeo Mn =', ChMnh)
    c3ent15.delete(0, 'end');                 c3ent15.insert (0, ChMnh) 
    
    #Diseño a cortante
    print(''); print('Diseño a cortante')

    Vcv = f_Vc(fiv, fc, b, d);                  print(" Vcv = %.1f" %Vcv)
    c4ent05.delete(0, 'end');                   c4ent05.insert (0, "{:.2f}".format(Vcv))
    ChVcv = f_ChVc(Vux,Vcv);                    print(' Chequeo cortante = ', ChVcv)
    c4ent06.delete(0, 'end');                   c4ent06.insert (0, ChVcv)

    Vch = f_Vc(fiv, fc, b, d);                  print(" Vch = %.1f" %Vch)
    c4ent55.delete(0, 'end');                   c4ent55.insert (0, "{:.2f}".format(Vch))
    ChVch = f_ChVc(Vux,Vch);                    print(' Chequeo cortante = ', ChVch)
    c4ent56.delete(0, 'end');                   c4ent56.insert (0, ChVch)
  

# Función guardar
def Guardar():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    book = load_workbook('D:\IEB_MUROS\PlantillaWallDesign.xlsx')
    sheet = book.active
    
    # Materiales
    fc = float (c1ent01.get())
    fy = float (c1ent02.get())
    b = float (c1ent03.get());                  sheet['D7'] = b; sheet['C35'] = b*10; sheet['D35'] = b*10
    h = float (c1ent04.get());                  sheet['D5'] = h
    r = float (c1ent05.get())
    fiv = float (c1ent06.get())
    pminv =  float (c1ent07.get())
    pminh =  float (c1ent08.get())
    d = h-r;                                    sheet['D6'] = d; sheet['C34'] = d*10; sheet['D34'] = d*10

    # Datos iniciales refuerzo vertical
    Mpy = float (c6ent01.get())
    Mny = float (c6ent02.get())
    Nbv = float (c2ent08.get())

    # Datos iniciales refuerzo horizontal
    Mpx = float (c6ent03.get())
    Mnx = float (c6ent04.get())
    Nbh = float (c3ent08.get())

    # Datos iniciales diseño cortante simple
    Vuy = float (c6ent05.get());                sheet['C36'] = "%.1f" %Vuy
    Vux = float (c6ent06.get());                sheet['D36'] = "%.1f" %Vux

    # Acero minimo
    Asminv = f_Asmin(pminv, b, h);              sheet['D13'] = "%.2f" %Asminv
    Asminh = f_Asmin(pminh, b, h);              sheet['D25'] = "%.2f" %Asminh

    #Diseño refuerzo vertical.
    print(''); print('Refuerzo vertical.')
    Muv = max (Mpy, Mny);                       sheet['D10'] = "%.1f" %Muv                      
    pcalv = f_pcal(fc, fy, Muv, b, d);          sheet['D11'] = "%.4f" %pcalv
    Ascalv = f_Ascal(pcalv, b, d);              sheet['D12'] = "%.2f" %Ascalv
    Asreqv = f_Asreq(Ascalv, Asminv);           sheet['D14'] = "%.2f" %Asreqv
    Asbv = f_Asb(Nbv);                         
    Sepv = f_sep(Asreqv, Asbv, b);              sheet['D16'] = "%.1f" %Sepv
    dbpv = f_dbp(Nbv);                          sheet['D15'] = dbpv                        
    Disv = f_dist(Asreqv, Asbv, b, dbpv);      
    Ascolv = f_Ascol(Asreqv, Asbv);            
    Mnv = f_Mn(fc, fy, Ascolv, b, d);           sheet['D18'] = "%.1f" %Mnv          
    ChMnv = f_ChMn(Muv,Mnv);                    sheet['D19'] = ChMnv

    #Diseño refuerzo horizontal.
    print(''); print('Refuerzo Horizontal')
    Muh = max (Mpx, Mnx);                       sheet['D22'] = "%.1f" %Muh
    pcalh = f_pcal(fc, fy, Muh, b, d);          sheet['D23'] = "%.4f" %pcalh
    Ascalh = f_Ascal(pcalh, b, d);              sheet['D24'] = "%.2f" %Ascalh
    Asreqh = f_Asreq(Ascalh, Asminh);           sheet['D26'] = "%.2f" %Asreqh
    Asbh = f_Asb(Nbh);                        
    Seph = f_sep(Asreqh, Asbh, b);              sheet['D28'] = "%.1f" %Seph
    dbph = f_dbp(Nbh);                          sheet['D27'] = dbph 
    Dish = f_dist(Asreqh, Asbh, b, dbph);     
    Ascolh = f_Ascol(Asreqh, Asbh);           
    Mnh = f_Mn(fc, fy, Ascolh, b, d);           sheet['D30'] = "%.1f" %Mnh 
    ChMnh = f_ChMn(Muh,Mnh);                    sheet['D31'] = ChMnh

    #Diseño a cortante
    print(''); print('Diseño a cortante')

    Vcv = f_Vc(fiv, fc, b, d);                   sheet['C37'] = "%.1f" %Vcv                  
    ChVcv = f_ChVc(Vux,Vcv);                     sheet['C38'] = ChVcv 

    Vch = f_Vc(fiv, fc, b, d);                   sheet['D37'] = "%.1f" %Vch 
    ChVch = f_ChVc(Vux,Vch);                     sheet['D38'] = ChVch 
    

    book.save ('D:\IEB_REPORTES\Wall Design.xlsx')

# Funcion borrar datos
def Borrar ():
    c1ent01.delete(0, 'end')
    c1ent02.delete(0, 'end')
    c1ent03.delete(0, 'end')
    c1ent04.delete(0, 'end')
    c1ent05.delete(0, 'end')
    c1ent06.delete(0, 'end')
    c1ent07.delete(0, 'end')
    c1ent08.delete(0, 'end')
    c1ent09.delete(0, 'end')
    c1ent10.delete(0, 'end')

    c2ent04.delete(0, 'end')
    c2ent05.delete(0, 'end')
    c2ent06.delete(0, 'end')
    c2ent07.delete(0, 'end')
    c2ent08.delete(0, 'end')
    c2ent09.delete(0, 'end')
    c2ent13.delete(0, 'end')
    c2ent14.delete(0, 'end')
    c2ent15.delete(0, 'end')

    c3ent04.delete(0, 'end')
    c3ent05.delete(0, 'end')
    c3ent06.delete(0, 'end')
    c3ent07.delete(0, 'end')
    c3ent08.delete(0, 'end')
    c3ent09.delete(0, 'end')
    c3ent13.delete(0, 'end')
    c3ent14.delete(0, 'end')
    c3ent15.delete(0, 'end')

    c4ent05.delete(0, 'end');       c4ent55.delete(0, 'end')
    c4ent06.delete(0, 'end');       c4ent56.delete(0, 'end')

    c6ent01.delete(0, 'end')
    c6ent02.delete(0, 'end')
    c6ent03.delete(0, 'end')
    c6ent04.delete(0, 'end')
    c6ent05.delete(0, 'end')
    c6ent06.delete(0, 'end')

# APLICACION TIPO ESCRITORIO
from tkinter import *
vent = Tk()
vent.geometry("760x900")
vent.title(" WALL DESING ")
vent.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico')

# Recuadro 1 Datos iniciales

rec1 = LabelFrame(vent, text = '  Datos iniciales.  ')
rec1.pack()
rec1.place(x=5, y=5, width=370, height=330)

c1tex01 = Label(rec1, text = "fc. Resistencia del concreto _ MPa"); c1tex01.pack()
c1tex01.place(x=10, y=10, width=210, height=20)
c1ent01 = Entry(rec1, justify=CENTER);                              c1ent01.place(x=230, y=10, width=80, height=20)

c1tex02 = Label(rec1, text = "fy. Fluencia del acero _ MPa");       c1tex02.pack()
c1tex02.place(x=10, y=40, width=210, height=20)
c1ent02 = Entry(rec1, justify=CENTER);                              c1ent02.place(x=230, y=40, width=80, height=20)

c1tex03 = Label(rec1, text = "b. Ancho de la sección _ cm");        c1tex03.pack()
c1tex03.place(x=10, y=70, width=210, height=20)
c1ent03 = Entry(rec1, justify=CENTER);                              c1ent03.place(x=230, y=70, width=80, height=20)

c1tex04 = Label(rec1, text = "h. Espesor del muro _ cm");           c1tex04.pack()
c1tex04.place(x=10, y=100, width=210, height=20)
c1ent04 = Entry(rec1, justify=CENTER);                              c1ent04.place(x=230, y=100, width=80, height=20)

c1tex05 = Label(rec1, text = "r. Recubrimiento _ cm");              c1tex05.pack()
c1tex05.place(x=10, y=130, width=210, height=20)
c1ent05 = Entry(rec1, justify=CENTER);                              c1ent05.place(x=230, y=130, width=80, height=20)

c1tex06 = Label(rec1, text = "øv. Coeficiente de fricción.");       c1tex06.pack()
c1tex06.place(x=10, y=160, width=210, height=20)
c1ent06 = Entry(rec1, justify=CENTER);                              c1ent06.place(x=230, y=160, width=80, height=20)

c1tex07 = Label(rec1, text = "pmínv. Cuantía mín vertical");        c1tex07.pack()
c1tex07.place(x=10, y=190, width=210, height=20)
c1ent07 = Entry(rec1, justify=CENTER);                              c1ent07.place(x=230, y=190, width=80, height=20)

c1tex08 = Label(rec1, text = "pmính. Cuantía mín horizontal");     c1tex08.pack()
c1tex08.place(x=10, y=220, width=210, height=20)
c1ent08 = Entry(rec1, justify=CENTER);                              c1ent08.place(x=230, y=220, width=80, height=20)

c1tex09 = Label(rec1, text = "Asminv. Acero mín vertical _ cm²");            c1tex09.pack()
c1tex09.place(x=10, y=250, width=210, height=20)
c1ent09 = Entry(rec1, justify=CENTER);                                       c1ent09.place(x=230, y=250, width=80, height=20); c1ent09.config(bg="#ecf0f1")

c1tex10 = Label(rec1, text = "Asminh. Acero mín horizontal _ cm²");          c1tex10.pack()
c1tex10.place(x=10, y=280, width=210, height=20)
c1ent10 = Entry(rec1, justify=CENTER);                                       c1ent10.place(x=230, y=280, width=80, height=20); c1ent10.config(bg="#ecf0f1")

# Recuadro 6 Momentos y Cortantes de diseño

rec6 = LabelFrame(vent, text = '  Momentos y fuerzas cortantes de diseño.  ')
rec6.pack()
rec6.place(x=385, y=5, width=370, height=210)

c6tex01 = Label(rec6, text = "Myy (+) Momento vertical positivo _ kN.cm");      c6tex01.pack()
c6tex01.place(x=10, y=10, width=250, height=20)
c6ent01 = Entry(rec6, justify=CENTER);                                          c6ent01.place(x=280, y=10, width=80, height=20)

c6tex02 = Label(rec6, text = "Myy (-) Momento vertical negativo _ kN.cm");      c6tex02.pack()
c6tex02.place(x=10, y=40, width=250, height=20)
c6ent02 = Entry(rec6, justify=CENTER);                                          c6ent02.place(x=280, y=40, width=80, height=20)

c6tex03 = Label(rec6, text = "Mxx (+) Momento horizontal positivo _ kN.cm");    c6tex03.pack()
c6tex03.place(x=10, y=70, width=250, height=20)
c6ent03 = Entry(rec6, justify=CENTER);                                          c6ent03.place(x=280, y=70, width=80, height=20)

c6tex04 = Label(rec6, text = "Mxx (-) Momento horizontal negativo _ kN.cm");    c6tex04.pack()
c6tex04.place(x=10, y=100, width=250, height=20)
c6ent04 = Entry(rec6, justify=CENTER);                                          c6ent04.place(x=280, y=100, width=80, height=20)

c6tex05 = Label(rec6, text = "Vyy Fuerza cortante vertical _ kN");              c6tex05.pack()
c6tex05.place(x=10, y=130, width=250, height=20)
c6ent05 = Entry(rec6, justify=CENTER);                                          c6ent05.place(x=280, y=130, width=80, height=20)

c6tex06 = Label(rec6, text = "Vxx Fuerza cortante horizontal _ kN");            c6tex06.pack()
c6tex06.place(x=10, y=160, width=250, height=20)
c6ent06 = Entry(rec6, justify=CENTER);                                          c6ent06.place(x=280, y=160, width=80, height=20)

# Recuadro 2 Refuerzo Vertical

rec2 = LabelFrame(vent, text = '  Diseño Refuerzo vertical.  ')
rec2.pack()
rec2.place(x=5, y=345, width=370, height=330)

c2tex01 = Label(rec2, text = "");                                   c2tex01.pack()
c2tex01.place(x=10, y=5, width=120, height=20)

c2tex04 = Label(rec2, text = "Momento de diseño _ kN.cm");          c2tex01.pack()
c2tex04.place(x=10, y=10, width=210, height=20)
c2ent04 = Entry(rec2, justify=CENTER);                              c2ent04.place(x=230, y=10, width=120, height=20);   c2ent04.config(bg="#ecf0f1")

c2tex05 = Label(rec2, text = "preq. Cuantía requerida.");            c2tex05.pack()
c2tex05.place(x=10, y=40, width=210, height=20)
c2ent05 = Entry(rec2, justify=CENTER);                              c2ent05.place(x=230, y=40, width=120, height=20); c2ent05.config(bg="#ecf0f1")

c2tex06 = Label(rec2, text = "Asreq. Acero requerido _ cm²");       c2tex06.pack()
c2tex06.place(x=10, y=70, width=210, height=20)
c2ent06 = Entry(rec2, justify=CENTER);                              c2ent06.place(x=230, y=70, width=120, height=20); c2ent06.config(bg="#ecf0f1")

c2tex07 = Label(rec2, text = "Ascol. Acero colocado _ cm²");        c2tex07.pack()
c2tex07.place(x=10, y=100, width=210, height=20)
c2ent07 = Entry(rec2, justify=CENTER);                              c2ent07.place(x=230, y=100, width=120, height=20); c2ent07.config(bg="#ecf0f1")

c2tex08 = Label(rec2, text = "N° de barra a usar.");                 c2tex08.pack()
c2tex08.place(x=10, y=130, width=210, height=20)
c2ent08 = Entry(rec2, justify=CENTER);                              c2ent08.place(x=230, y=130, width=120, height=20)

c2tex09 = Label(rec2, text = "Distribución del refuerzo.");          c2tex09.pack()
c2tex09.place(x=10, y=160, width=210, height=20)
c2ent09 = Entry(rec2, justify=CENTER);                              c2ent09.place(x=230, y=160, width=120, height=20); c2ent09.config(bg="#ecf0f1")

c2tex10 = Label(rec2, text = "Chequeo momento nominal.");            c2tex10.pack()
c2tex10.place(x=10, y=190, width=210, height=20)

c2tex13 = Label(rec2, text = "Ascol. Acero colocado real _ cm²");    c2tex13.pack()
c2tex13.place(x=10, y=220, width=210, height=20)
c2ent13 = Entry(rec2, justify=CENTER);                              c2ent13.place(x=230, y=220, width=120, height=20); c2ent13.config(bg="#ecf0f1")

c2tex14 = Label(rec2, text = "øMn. Momento nominal _ kN.cm");       c2tex14.pack()
c2tex14.place(x=10, y=250, width=210, height=20)
c2ent14 = Entry(rec2, justify=CENTER);                              c2ent14.place(x=230, y=250, width=120, height=20); c2ent14.config(bg="#ecf0f1")

c2tex15 = Label(rec2, text = "Mu. < øMn.");                         c2tex15.pack()
c2tex15.place(x=10, y=280, width=210, height=20)
c2ent15 = Entry(rec2, justify=CENTER);                              c2ent15.place(x=230, y=280, width=120, height=20); c2ent15.config(bg="#ecf0f1")

# Recuadro 3. Refuerzo horizontal

rec3 = LabelFrame(vent, text = '  Diseño refuerzo horizontal.  ')
rec3.pack()
rec3.place(x=385, y=345, width=370, height=330)

c3tex01 = Label(rec3, text = "");                                    c3tex01.pack()
c3tex01.place(x=10, y=5, width=120, height=20)

c3tex04 = Label(rec3, text = "Momento de diseño _ kN.cm");          c3tex01.pack()
c3tex04.place(x=10, y=10, width=210, height=20)
c3ent04 = Entry(rec3, justify=CENTER);                              c3ent04.place(x=230, y=10, width=120, height=20);   c3ent04.config(bg="#ecf0f1")

c3tex05 = Label(rec3, text = "preq. Cuantía requerida.");            c3tex05.pack()
c3tex05.place(x=10, y=40, width=210, height=20)
c3ent05 = Entry(rec3, justify=CENTER);                              c3ent05.place(x=230, y=40, width=120, height=20); c3ent05.config(bg="#ecf0f1")

c3tex06 = Label(rec3, text = "Asreq. Acero requerido _ cm²");       c3tex06.pack()
c3tex06.place(x=10, y=70, width=210, height=20)
c3ent06 = Entry(rec3, justify=CENTER);                              c3ent06.place(x=230, y=70, width=120, height=20); c3ent06.config(bg="#ecf0f1")

c3tex07 = Label(rec3, text = "Ascol. Acero colocado _ cm²");        c3tex07.pack()
c3tex07.place(x=10, y=100, width=210, height=20)
c3ent07 = Entry(rec3, justify=CENTER);                              c3ent07.place(x=230, y=100, width=120, height=20); c3ent07.config(bg="#ecf0f1")

c3tex08 = Label(rec3, text = "N° de barra a usar.");                 c3tex08.pack()
c3tex08.place(x=10, y=130, width=210, height=20)
c3ent08 = Entry(rec3, justify=CENTER);                              c3ent08.place(x=230, y=130, width=120, height=20)

c3tex09 = Label(rec3, text = "Distribución del refuerzo.");          c3tex09.pack()
c3tex09.place(x=10, y=160, width=210, height=20)
c3ent09 = Entry(rec3, justify=CENTER);                              c3ent09.place(x=230, y=160, width=120, height=20); c3ent09.config(bg="#ecf0f1")

c3tex10 = Label(rec3, text = "Chequeo momento nominal.");            c3tex10.pack()
c3tex10.place(x=10, y=190, width=210, height=20)

c3tex13 = Label(rec3, text = "Ascol. Acero colocado real _ cm²");   c3tex13.pack()
c3tex13.place(x=10, y=220, width=210, height=20)
c3ent13 = Entry(rec3, justify=CENTER);                              c3ent13.place(x=230, y=220, width=120, height=20); c3ent13.config(bg="#ecf0f1")

c3tex14 = Label(rec3, text = "øMn. Momento nominal _ kN.cm");       c3tex14.pack()
c3tex14.place(x=10, y=250, width=210, height=20)
c3ent14 = Entry(rec3, justify=CENTER);                              c3ent14.place(x=230, y=250, width=120, height=20); c3ent14.config(bg="#ecf0f1")

c3tex15 = Label(rec3, text = "Mu. < øMn.");                         c3tex15.pack()
c3tex15.place(x=10, y=280, width=210, height=20)
c3ent15 = Entry(rec3, justify=CENTER);                              c3ent15.place(x=230, y=280, width=120, height=20); c3ent15.config(bg="#ecf0f1")

# Recuadro 4 Cortante

rec4 = LabelFrame(vent, text = '  Cortante simple.  ')
rec4.pack()
rec4.place(x=135, y=685, width=490, height=120)

c4tex01 = Label(rec4, text = "Chequeo a cortante simple.");            c4tex01.pack()
c4tex01.place(x=10, y=10, width=210, height=20)

c4tex02 = Label(rec4, text = "Vc _ Vertical", font='Helvetica 7 bold');  c4tex02.pack()
c4tex02.place(x=230, y=10, width=120, height=20)

c4tex03 = Label(rec4, text = "Vc _ Horizontal", font='Helvetica 7 bold'); c4tex03.pack()
c4tex03.place(x=360, y=10, width=120, height=20)

c4tex05 = Label(rec4, text = "øVc. Resitencia del concreto _ kN");  c4tex05.pack()
c4tex05.place(x=10, y=40, width=210, height=20)
c4ent05 = Entry(rec4, justify=CENTER);                              c4ent05.place(x=230, y=40, width=120, height=20); c4ent05.config(bg="#ecf0f1")
c4ent55 = Entry(rec4, justify=CENTER);                              c4ent55.place(x=360, y=40, width=120, height=20); c4ent55.config(bg="#ecf0f1")

c4tex06 = Label(rec4, text = "Vu. < øVc.");                         c4tex06.pack()
c4tex06.place(x=10, y=70, width=210, height=20)
c4ent06 = Entry(rec4, justify=CENTER);                              c4ent06.place(x=230, y=70, width=120, height=20); c4ent06.config(bg="#ecf0f1")
c4ent56 = Entry(rec4, justify=CENTER);                              c4ent56.place(x=360, y=70, width=120, height=20); c4ent56.config(bg="#ecf0f1")

# Recuadro 5 botones

rec5 = LabelFrame(vent, text = '  Opciones.  ')
rec5.pack()
rec5.place(x=135, y=815, width=490, height=60)

bot1 = Button(rec5, text = 'Calcular', font='Helvetica 8 bold', command=Calcular); bot1.pack()
bot1.place(x=47.5, y=10, width=100, height=20)

bot2 = Button(rec5, text = 'Borrar', font='Helvetica 8 bold', command=Borrar); bot2.pack()
bot2.place(x=195, y=10, width=100, height=20)

bot3 = Button(rec5, text = 'Guardar .xls', font='Helvetica 8 bold', command=Guardar); bot3.pack()
bot3.place(x=342.5, y=10, width=100, height=20)

label = Label(vent, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7'); label.pack()
label.place(x=135, y=885, width=490, height=10)

vent.mainloop()

